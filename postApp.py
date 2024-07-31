from flask import Flask, jsonify, request
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment, NamedStyle
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import Spacer
import subprocess

TEN_SHOP = "SHOP NGUYEN TRUNG" # Tên shop
san_pham_dict = {}


def doc_file_excel(file_path):
    global san_pham_dict
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua dòng tiêu đề
            ma_sp = row[0]
            gia_ban_str = str(row[1])

            # Chuyển đổi giá bán từ chuỗi sang số, loại bỏ dấu phẩy
            gia_ban_str = gia_ban_str.replace(',', '')
            gia_ban = int(gia_ban_str)

            # Thêm sản phẩm vào từ điển
            san_pham_dict[ma_sp] = gia_ban

    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")

    return san_pham_dict

# Sử dụng hàm
file_path = "HangHoa.xlsx"  # Thay thế bằng đường dẫn thực tế
san_pham_dict = doc_file_excel(file_path)

app = Flask(__name__)
@app.route('/api', methods=['POST'])
def tao_hoa_don():
    global san_pham_dict
    global centered_style_new
    ngay_gio_hien_tai = datetime.now()
    ten_file = ngay_gio_hien_tai.strftime("%d_%m_%Y") + ".xlsx"
    hoa_don_pdf_file = ngay_gio_hien_tai.strftime("%d_%m_%Y_%H_%M_%S") + ".pdf"
    ten_sheet = ngay_gio_hien_tai.strftime("%H") + "H"

    try:
        workbook = openpyxl.load_workbook(ten_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    if "centered_style" in workbook.named_styles:
        centered_style_new = workbook.named_styles[len(workbook.named_styles) -1]
    else:
        # Tạo kiểu định dạng căn giữa
        centered_style_new = NamedStyle(name="centered_style")
        centered_style_new.alignment = Alignment(horizontal='center', vertical='center')

    check = False
    if ten_sheet not in workbook.sheetnames:
        workbook.create_sheet(ten_sheet)
        check = True
    sheet = workbook[ten_sheet]
    if check:
        for col_idx in range(1, 9):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].auto_size = True
    sheet.cell(row=1, column=1, value="Thời gian").style = centered_style_new
    sheet.cell(row=1, column=2, value="Số tiền thanh toán").style = centered_style_new
    sheet.cell(row=1, column=3, value="Tiền KM").style = centered_style_new
    sheet.cell(row=1, column=4, value="Mua/Trả").style = centered_style_new
    sheet.cell(row=1, column=5, value="Mã SP").style = centered_style_new
    sheet.cell(row=1, column=6, value="Tên SP").style = centered_style_new
    sheet.cell(row=1, column=7, value="Số lượng").style = centered_style_new
    sheet.cell(row=1, column=8, value="Đơn giá").style = centered_style_new

    try:
        data = request.get_json()
        san_pham_mua = data['san_pham_mua']
        so_tien_giam = data['so_tien_giam']
        san_pham_tra = data.get('san_pham_tra', {})  # Nếu không có sản phẩm trả, mặc định là từ điển rỗng

        tong_tien = 0
        for ma_sp, so_luong in san_pham_mua.items():
            gia_ban = san_pham_dict.get(int(ma_sp))
            if gia_ban:
                tong_tien += gia_ban * so_luong

        for ma_sp, so_luong in san_pham_tra.items():
            gia_ban = san_pham_dict.get(int(ma_sp))
            if gia_ban:
                tong_tien -= gia_ban * so_luong

        tong_sp = len(san_pham_mua) + len(san_pham_tra)

        # Thêm dòng mới
        row_index = sheet.max_row + 1
        sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index+tong_sp -1, end_column=1)
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index+tong_sp -1, end_column=2)
        sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index+tong_sp -1, end_column=3)
        sheet.merge_cells(start_row=row_index, start_column=4, end_row=row_index+len(san_pham_mua) -1, end_column=4)
        sheet.merge_cells(start_row=row_index+len(san_pham_mua), start_column=4, end_row=row_index+tong_sp -1, end_column=4)
        sheet.cell(row=row_index, column=1, value=ngay_gio_hien_tai.strftime("%H:%M:%S")).style = centered_style_new
        tong_tien_formatted = "{:,.0f} VND".format(tong_tien-so_tien_giam)
        sheet.cell(row=row_index, column=2, value=tong_tien_formatted).style = centered_style_new
        so_tien_giam_formatted = "Note: Đã giảm {:,.0f} VND".format(so_tien_giam)
        sheet.cell(row=row_index, column=3, value=so_tien_giam_formatted).style = centered_style_new

        # # Cột (Mua/Trả)
        sheet.cell(row=row_index, column=4, value="Mua").style = centered_style_new
        sheet.cell(row=row_index+len(san_pham_mua), column=4, value="Trả").style = centered_style_new

        # Tiêu đề bảng
        data = [["Ten SP", "SL", "Gia", "Thanh tien"]]
        # # Các cột chi tiết sản phẩm (Số lượng, Đơn giá)
        for ma_sp, so_luong in san_pham_mua.items():
            sheet.cell(row=row_index, column=5, value=so_luong).style = centered_style_new
            sheet.cell(row=row_index, column=6, value="SHOP NGUYEN TRUNG").style = centered_style_new
            sheet.cell(row=row_index, column=7, value=so_luong).style = centered_style_new
            
            gia_ban = san_pham_dict.get(int(ma_sp))
            gia_ban_formatted = "{:,.0f} VND".format(gia_ban)
            sheet.cell(row=row_index, column=8, value=gia_ban_formatted).style = centered_style_new
            row_index += 1
            data.append(["SHOP NGUYEN TRUNG", ma_sp, "{:,.0f}".format(gia_ban), "{:,.0f}".format(so_luong*gia_ban)])
        data.append([])
        for ma_sp, so_luong in san_pham_tra.items():
            sheet.cell(row=row_index, column=5, value=so_luong).style = centered_style_new
            sheet.cell(row=row_index, column=6, value="SHOP NGUYEN TRUNG").style = centered_style_new
            sheet.cell(row=row_index, column=7, value=so_luong).style = centered_style_new
            
            gia_ban = int(san_pham_dict.get(int(ma_sp)))
            gia_ban_formatted = "{:,.0f} VND".format(gia_ban)
            sheet.cell(row=row_index, column=8, value=gia_ban_formatted).style = centered_style_new
            row_index += 1
            data.append([ "SHOP NGUYEN TRUNG", ma_sp, "- {:,.0f}".format(gia_ban), "- {:,.0f}".format(so_luong*gia_ban)])
        data.append([])
        data.append(["", "", "Tong:", "{:,.0f}".format(tong_tien)])
        data.append(["", "", "Giam gia:", "{:,.0f}".format(so_tien_giam)])
        data.append(["", "", "Thanh tien:", "{:,.0f}".format(tong_tien-so_tien_giam)])
        workbook.save(ten_file)
        # Tính toán chiều cao ước lượng của hóa đơn
        row_height = 20  # Chiều cao ước tính của mỗi dòng (điều chỉnh nếu cần)
        header_height = 60  # Chiều cao của tiêu đề và ngày giờ
        footer_height = 60  # Chiều cao của tổng, giảm giá, thành tiền và note
        table_height = (len(san_pham_mua) + len(san_pham_tra) + 3 + 2) * row_height  # Chiều cao của bảng
        estimated_height = header_height + table_height + footer_height + 40
        doc = SimpleDocTemplate(
            hoa_don_pdf_file, 
            pagesize=(80 * mm, estimated_height),
            leftMargin=5,  #   lề trái
            rightMargin=5,  #   lề phải
            topMargin=10,   #   lề trên
            bottomMargin=10  #   lề dưới
        )

        # Tạo nội dung hóa đơn
        story = []

        # Tên shop (in đậm, size lớn, căn giữa)
        title_data = [["SHOP NGUYEN TRUNG"]]
        title_table = Table(title_data, colWidths=[80 * mm])
        title_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 14)  # Điều chỉnh kích thước chữ ở đây
        ]))
        story.append(title_table)

        # Ngày giờ (căn giữa)
        time_data = [[ngay_gio_hien_tai.strftime("%d/%m/%Y %H:%M:%S")]]
        time_data = Table(time_data, colWidths=[80 * mm])
        time_data.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10)  # Điều chỉnh kích thước chữ ở đây
        ]))
        story.append(time_data)

        story.append(Spacer(1, 10))

        # Tạo bảng
        table = Table(data)
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('SIZE', (0, 0), (-1, -1), 8)
        ]))
        story.append(table)
        story.append(Spacer(1, 10))

        # Ngày giờ (căn giữa)
        note_Data = [["Thank you\nHang moi duoc doi tra trong vong 7 ngay"]]
        note_Data = Table(note_Data, colWidths=[80 * mm])
        note_Data.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 10)  # Điều chỉnh kích thước chữ ở đây
        ]))
        story.append(note_Data)

        # Build PDF
        # doc.build(story)
        
        try:
            subprocess.Popen(["start", "", hoa_don_pdf_file], shell=True)  # Mở file PDF trong trình xem mặc định và in
        except Exception as e:
            return jsonify({'error': f'Lỗi khi in: {e}'}), 500
        return jsonify({'tong_tien': tong_tien})
    except KeyError:
        return jsonify({'error': 'Dữ liệu không hợp lệ'}), 400

@app.route('/')
def hello():
    return "Xin chào! Đây là máy chủ web của Anh Hiếu! Heheheeeeeeeee"


@app.route('/api/<int:product_id>', methods=['GET'])
def getPrice(product_id):
    return str(san_pham_dict.get(product_id))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000) 